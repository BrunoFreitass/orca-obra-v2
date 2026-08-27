"""Testes de core/sinapi_import.py -- leitura de planilhas oficiais do
SINAPI, incluindo o formato de pacote nacional consolidado (cabecalho
de duas linhas nas abas de composicao, codigo real so na aba
Analitico, varias abas com o mesmo formato para regimes diferentes de
encargos)."""
import pytest
from openpyxl import Workbook

from core import sinapi_import as si


def _sheet_insumos_simples(wb, titulo="ISD", regime="SEM DESONERAÇÃO", codigo=1379, preco_rr=1.62):
    """Formato de cabecalho de uma linha so (insumos): a UF aparece na
    mesma linha que 'Codigo'."""
    ws = wb.create_sheet(titulo)
    ws.append(["RELATÓRIO DE PREÇOS DE INSUMOS - ENCARGOS SOCIAIS " + regime])
    ws.append([])
    ws.append(["Classificação", "Código do\nInsumo", "Descrição do Insumo", "Unidade", "AC", "RR"])
    ws.append(["MATERIAL", codigo, "CIMENTO PORTLAND COMPOSTO CP II-32", "KG", 1.50, preco_rr])
    return ws


def _sheet_composicoes_duas_linhas(wb, titulo="CSD", regime="SEM DESONERAÇÃO", custo_rr=95.4, codigo_zerado=True):
    """Formato de cabecalho de duas linhas (composicoes): a UF fica numa
    linha acima da linha com 'Codigo da Composicao'/'Custo (R$)',
    alinhada na mesma coluna -- e o codigo real vem sempre 0 nesse
    pacote, como no arquivo nacional consolidado."""
    ws = wb.create_sheet(titulo)
    ws.append(["RELATÓRIO DE CUSTOS DE COMPOSIÇÕES - ENCARGOS SOCIAIS " + regime])
    ws.append([])
    ws.append([None, None, None, None, "AC", None, "RR"])
    ws.append(["Grupo", "Código da\nComposição", "Descrição", "Unidade", "Custo (R$)", "%AS", "Custo (R$)", "%AS"])
    codigo = 0 if codigo_zerado else 103361
    ws.append([
        "Alvenaria de Vedação", codigo,
        "ALVENARIA DE VEDAÇÃO DE BLOCOS CERÂMICOS FURADOS NA HORIZONTAL DE 14X19X29 CM "
        "(ESPESSURA 14 CM) E ARGAMASSA DE ASSENTAMENTO COM PREPARO MANUAL. AF_12/2021",
        "M2", 80.0, 0, custo_rr, 0,
    ])
    return ws


def _sheet_analitico(wb):
    ws = wb.create_sheet("Analítico")
    ws.append(["RELATÓRIO ANALÍTICO DE COMPOSIÇÕES"])
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append(["Grupo", "Código da\nComposição", "Tipo Item", "Código do\nItem", "Descrição", "Unidade", "Coeficiente", "Situação"])
    # linha-resumo da composicao (Tipo Item vazio) -- e' esta que traz o codigo real
    ws.append([
        "Alvenaria de Vedação", 103361, None, None,
        "ALVENARIA DE VEDAÇÃO DE BLOCOS CERÂMICOS FURADOS NA HORIZONTAL DE 14X19X29 CM "
        "(ESPESSURA 14 CM) E ARGAMASSA DE ASSENTAMENTO COM PREPARO MANUAL. AF_12/2021",
        "M2", None, "COM CUSTO",
    ])
    # sub-item (Tipo Item preenchido) -- nao deve ser confundido com a linha-resumo
    ws.append([
        "Alvenaria de Vedação", 103361, "INSUMO", 34353,
        "ARGAMASSA COLANTE AC II", "KG", 12.5, "COM PREÇO",
    ])
    return ws


def _workbook_basico(tmp_path, nome="sinapi.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)
    return wb, tmp_path / nome


class TestCabecalhoUmaLinha:
    def test_le_insumo_com_preco_rr(self, tmp_path):
        wb, caminho = _workbook_basico(tmp_path)
        _sheet_insumos_simples(wb, preco_rr=1.62)
        wb.save(caminho)

        encontrados = si.ler_planilha_sinapi(caminho)

        assert encontrados["1379"]["preco"] == 1.62
        assert encontrados["1379"]["unidade"] == "KG"


class TestCabecalhoDuasLinhas:
    def test_sem_aba_analitico_no_codigo_zerado_e_ignorado(self, tmp_path):
        """Sem a aba Analitico pra resolver o codigo, uma composicao com
        codigo 0 nao deve ser incluida (0 nunca e' um codigo SINAPI
        valido -- incluir geraria colisao silenciosa entre composicoes
        diferentes)."""
        wb, caminho = _workbook_basico(tmp_path)
        _sheet_composicoes_duas_linhas(wb, custo_rr=95.4, codigo_zerado=True)
        wb.save(caminho)

        encontrados = si.ler_planilha_sinapi(caminho)

        assert "0" not in encontrados
        assert encontrados == {}

    def test_com_aba_analitico_resolve_codigo_real(self, tmp_path):
        wb, caminho = _workbook_basico(tmp_path)
        _sheet_composicoes_duas_linhas(wb, custo_rr=95.4, codigo_zerado=True)
        _sheet_analitico(wb)
        wb.save(caminho)

        encontrados = si.ler_planilha_sinapi(caminho)

        assert encontrados["103361"]["preco"] == 95.4
        assert encontrados["103361"]["unidade"] == "M2"

    def test_codigo_ja_populado_dispensa_aba_analitico(self, tmp_path):
        wb, caminho = _workbook_basico(tmp_path)
        _sheet_composicoes_duas_linhas(wb, custo_rr=95.4, codigo_zerado=False)
        wb.save(caminho)

        encontrados = si.ler_planilha_sinapi(caminho)

        assert encontrados["103361"]["preco"] == 95.4


class TestFiltroRegimeDesoneracao:
    def test_ignora_aba_com_desoneracao_quando_ha_sem_desoneracao(self, tmp_path):
        """Pacote nacional traz ISD (sem desoneracao) e ICD (com
        desoneracao) para o mesmo codigo -- so o preco 'sem
        desoneracao' deve ser usado, nunca sobrescrito pelo outro
        regime."""
        wb, caminho = _workbook_basico(tmp_path)
        _sheet_insumos_simples(wb, titulo="ISD", regime="SEM DESONERAÇÃO", codigo=1379, preco_rr=1.62)
        _sheet_insumos_simples(wb, titulo="ICD", regime="COM DESONERAÇÃO", codigo=1379, preco_rr=999.99)
        wb.save(caminho)

        encontrados = si.ler_planilha_sinapi(caminho)

        assert encontrados["1379"]["preco"] == 1.62

    def test_sem_nenhuma_aba_sem_desoneracao_processa_tudo(self, tmp_path):
        """Formato mais simples (relatorio so de RR, sem variantes de
        regime) nao deve ser quebrado pelo filtro -- se nenhuma aba
        bater com 'sem desoneracao', volta a processar todas."""
        wb, caminho = _workbook_basico(tmp_path)
        _sheet_insumos_simples(wb, titulo="Insumos", regime="", codigo=1379, preco_rr=1.62)
        wb.save(caminho)

        encontrados = si.ler_planilha_sinapi(caminho)

        assert encontrados["1379"]["preco"] == 1.62


class TestPrecoZerado:
    def test_composicao_sem_custo_para_rr_e_ignorada(self, tmp_path):
        """'Custo zerado' no relatorio do SINAPI indica que a UF nao
        tem coleta para aquela composicao -- 0 nao e' um preco real e
        nao deve ser aceito."""
        wb, caminho = _workbook_basico(tmp_path)
        _sheet_composicoes_duas_linhas(wb, custo_rr=0, codigo_zerado=False)
        wb.save(caminho)

        encontrados = si.ler_planilha_sinapi(caminho)

        assert "103361" not in encontrados


class TestImportarCompensaFatorRegional:
    """core/calculator.py multiplica TODO preco por FATOR_REGIONAL_RR na
    hora do calculo -- correto pros valores padrao de coeficientes.py
    (medias nacionais), mas os valores que vem do SINAPI ja sao a
    coluna RR (ja regionais). importar() precisa gravar o valor
    dividido por FATOR_REGIONAL_RR pra nao aplicar o fator 2x."""

    def test_cimento_grava_valor_dividido_pelo_fator_regional(self, tmp_path):
        from core.coeficientes import FATOR_REGIONAL_RR

        wb = Workbook()
        wb.remove(wb.active)
        _sheet_insumos_simples(wb, titulo="ISD", codigo=1379, preco_rr=1.62)
        caminho = tmp_path / "sinapi.xlsx"
        wb.save(caminho)

        precos, _avisos, _mes = si.importar([caminho], mes_referencia="2026-07")

        # 1.62 (preco bruto) x 50 (fator_conversao do cimento, kg -> saco
        # de 50kg) / 1.070 (fator regional, compensado na gravacao).
        # Tolerancia mais larga por causa do arredondamento pra 2 casas
        # decimais na gravacao (round(valor_final, 2) em importar()).
        esperado = (1.62 * 50) / FATOR_REGIONAL_RR.valor
        assert precos["cimento"]["valor"] == pytest.approx(esperado, abs=0.01)
        # conferencia: multiplicar de volta pelo fator regional (como
        # calculator.py faz) tem que devolver o preco real da planilha
        assert precos["cimento"]["valor"] * FATOR_REGIONAL_RR.valor == pytest.approx(81.0, abs=0.01)

    def test_aco_grava_valor_sem_compensar_fator_regional(self, tmp_path):
        # "aco" e' excecao: seu consumo em calculator.py nunca aplica o
        # fator regional (PRECO_ACO_RR, o padrao dela, ja era um valor
        # de RR antes deste importador existir) -- dividir aqui deixaria
        # o preco final baixo demais.
        wb = Workbook()
        wb.remove(wb.active)
        _sheet_insumos_simples(wb, titulo="ISD", codigo=32, preco_rr=9.47)
        caminho = tmp_path / "sinapi.xlsx"
        wb.save(caminho)

        precos, _avisos, _mes = si.importar([caminho], mes_referencia="2026-07")

        assert precos["aco"]["valor"] == pytest.approx(9.47, abs=0.001)
