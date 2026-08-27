from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def gerar_excel(dados_orcamento, output_path, bdi_percentual=0):
    xlsx_path = output_path.replace(".csv", ".xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Orçamento"

    fonte_padrao = "Arial"
    azul_cabecalho = "1F4E78"
    azul_secao = "4472A8"
    verde_venda = "2E7D32"
    cinza_claro = "F2F2F2"

    # --- CABECALHO PRINCIPAL ---
    # Aplica fill em TODAS as celulas ANTES do merge (openpyxl exige isso)
    fill_cabecalho = PatternFill(start_color=azul_cabecalho, end_color=azul_cabecalho, fill_type="solid")
    font_cabecalho = Font(name=fonte_padrao, size=14, bold=True, color="FFFFFF")
    align_cabecalho = Alignment(horizontal="center", vertical="center")
    for col in range(1, 5):
        c = ws.cell(row=1, column=col)
        c.fill = fill_cabecalho
        c.font = font_cabecalho
        c.alignment = align_cabecalho
    ws.merge_cells("A1:D1")
    ws["A1"] = "OrçaObra AI — Orçamento Estimado de Obra"
    ws.row_dimensions[1].height = 26

    # --- CABECALHO DA TABELA ---
    linha_cabecalho = 3
    colunas = ["Item", "Quantidade", "Preço Unitário (R$)", "Total (R$)"]
    fill_header = PatternFill(start_color=azul_cabecalho, end_color=azul_cabecalho, fill_type="solid")
    font_header = Font(name=fonte_padrao, size=11, bold=True, color="FFFFFF")
    align_header = Alignment(horizontal="center", vertical="center")
    for col_idx, titulo in enumerate(colunas, start=1):
        celula = ws.cell(row=linha_cabecalho, column=col_idx, value=titulo)
        celula.font = font_header
        celula.fill = fill_header
        celula.alignment = align_header

    borda_fina = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    grupos = {}
    ordem_grupos = []
    for item in dados_orcamento:
        tipo = item.get("Tipo", "Itens")
        if tipo not in grupos:
            grupos[tipo] = []
            ordem_grupos.append(tipo)
        grupos[tipo].append(item)

    linha_atual = linha_cabecalho + 1
    linhas_subtotal = []

    for tipo in ordem_grupos:
        itens_grupo = grupos[tipo]

        # --- TITULO DA SECAO (ex: MATERIAL) ---
        # Aplica fill em TODAS as celulas ANTES do merge
        fill_secao = PatternFill(start_color=azul_secao, end_color=azul_secao, fill_type="solid")
        font_secao = Font(name=fonte_padrao, size=11, bold=True, color="FFFFFF")
        align_secao = Alignment(horizontal="left", vertical="center", indent=1)
        for col in range(1, 5):
            c = ws.cell(row=linha_atual, column=col)
            c.fill = fill_secao
            c.font = font_secao
            c.alignment = align_secao
        ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
        ws.cell(row=linha_atual, column=1, value=tipo.upper())
        linha_atual += 1

        for i, item in enumerate(itens_grupo):
            zebra = cinza_claro if i % 2 == 1 else "FFFFFF"
            fill_zebra = PatternFill(start_color=zebra, end_color=zebra, fill_type="solid")

            total_calculado = round(item["Quantidade"] * item["Preco_Unit"], 2)

            ws.cell(row=linha_atual, column=1, value=item["Material"])
            ws.cell(row=linha_atual, column=2, value=item["Quantidade"])
            ws.cell(row=linha_atual, column=3, value=item["Preco_Unit"])
            ws.cell(row=linha_atual, column=4, value=total_calculado)

            for col_idx in range(1, 5):
                celula = ws.cell(row=linha_atual, column=col_idx)
                celula.font = Font(name=fonte_padrao, size=10, color="000000")
                celula.fill = fill_zebra
                celula.border = borda_fina
                if col_idx in (3, 4):
                    celula.number_format = 'R$ #,##0.00'
                if col_idx == 2:
                    celula.alignment = Alignment(horizontal="center")

            linha_atual += 1


        subtotal_valor = round(sum(it["Total"] for it in itens_grupo), 2)

        # --- SUBTOTAL ---
        ws.merge_cells(f"A{linha_atual}:C{linha_atual}")
        font_sub = Font(name=fonte_padrao, size=10, bold=True, italic=True, color="000000")
        align_sub = Alignment(horizontal="right", vertical="center")
        for col in range(1, 4):
            c = ws.cell(row=linha_atual, column=col)
            c.font = font_sub
            c.alignment = align_sub
        ws.cell(row=linha_atual, column=1, value=f"Subtotal — {tipo}")
        celula_subtotal = ws.cell(row=linha_atual, column=4, value=subtotal_valor)
        celula_subtotal.font = Font(name=fonte_padrao, size=10, bold=True, color="000000")
        celula_subtotal.number_format = 'R$ #,##0.00'
        linhas_subtotal.append(linha_atual)
        linha_atual += 2

    # --- CUSTO DIRETO ---
    linha_custo_direto = linha_atual
    label_custo = "CUSTO DIRETO" if bdi_percentual else "TOTAL ESTIMADO"
    # Aplica fill em TODAS as celulas ANTES do merge
    fill_custo = PatternFill(start_color=azul_cabecalho, end_color=azul_cabecalho, fill_type="solid")
    font_custo = Font(name=fonte_padrao, size=12, bold=True, color="FFFFFF")
    align_custo = Alignment(horizontal="right", vertical="center")
    for col in range(1, 4):
        c = ws.cell(row=linha_custo_direto, column=col)
        c.fill = fill_custo
        c.font = font_custo
        c.alignment = align_custo
    ws.merge_cells(f"A{linha_custo_direto}:C{linha_custo_direto}")
    ws.cell(row=linha_custo_direto, column=1, value=label_custo)

    custo_direto_valor = round(sum(item["Total"] for item in dados_orcamento), 2)

    celula_custo_direto = ws.cell(row=linha_custo_direto, column=4, value=custo_direto_valor)
    celula_custo_direto.font = font_custo
    celula_custo_direto.fill = fill_custo
    celula_custo_direto.number_format = 'R$ #,##0.00'
    ws.row_dimensions[linha_custo_direto].height = 22

    linha_preco_venda = linha_custo_direto

    if bdi_percentual:
        # --- BDI ---
        linha_bdi = linha_custo_direto + 1
        ws.merge_cells(f"A{linha_bdi}:C{linha_bdi}")
        font_bdi = Font(name=fonte_padrao, size=10, italic=True, color="000000")
        align_bdi = Alignment(horizontal="right", vertical="center")
        for col in range(1, 4):
            c = ws.cell(row=linha_bdi, column=col)
            c.font = font_bdi
            c.alignment = align_bdi
        ws.cell(row=linha_bdi, column=1,
                value=f"BDI ({bdi_percentual:g}%) — administração, lucro, impostos e imprevistos")
        # Mesma fórmula de core/orcamento_service.py::calcular_custo_e_preco --
        # bdi_valor é derivado por subtração (não recalculado a partir do
        # percentual) pra nunca divergir do Preço de Venda exibido logo
        # abaixo, nem do valor salvo no histórico.
        preco_venda_valor = round(custo_direto_valor * (1 + bdi_percentual / 100), 2)
        bdi_valor = round(preco_venda_valor - custo_direto_valor, 2)
        celula_bdi_valor = ws.cell(row=linha_bdi, column=4, value=bdi_valor)
        celula_bdi_valor.font = Font(name=fonte_padrao, size=10, color="000000")
        celula_bdi_valor.number_format = 'R$ #,##0.00'

        # --- PRECO DE VENDA ---
        linha_preco_venda = linha_bdi + 1
        # Aplica fill em TODAS as celulas ANTES do merge
        fill_venda = PatternFill(start_color=verde_venda, end_color=verde_venda, fill_type="solid")
        font_venda = Font(name=fonte_padrao, size=12, bold=True, color="FFFFFF")
        align_venda = Alignment(horizontal="right", vertical="center")
        for col in range(1, 4):
            c = ws.cell(row=linha_preco_venda, column=col)
            c.fill = fill_venda
            c.font = font_venda
            c.alignment = align_venda
        ws.merge_cells(f"A{linha_preco_venda}:C{linha_preco_venda}")
        ws.cell(row=linha_preco_venda, column=1, value="PREÇO DE VENDA")

        celula_venda_valor = ws.cell(row=linha_preco_venda, column=4, value=preco_venda_valor)
        celula_venda_valor.font = font_venda
        celula_venda_valor.fill = fill_venda
        celula_venda_valor.number_format = 'R$ #,##0.00'
        ws.row_dimensions[linha_preco_venda].height = 22

    # --- LARGURAS ---
    larguras = {1: 34, 2: 14, 3: 20, 4: 18}
    for col_idx, largura in larguras.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    # --- RODAPE ---
    linha_rodape = linha_preco_venda + 2
    ws.merge_cells(f"A{linha_rodape}:D{linha_rodape}")
    font_rodape = Font(name=fonte_padrao, size=8, italic=True, color="666666")
    align_rodape = Alignment(horizontal="left", vertical="center")
    for col in range(1, 5):
        c = ws.cell(row=linha_rodape, column=col)
        c.font = font_rodape
        c.alignment = align_rodape
    ws.cell(row=linha_rodape, column=1,
            value="Nota: os totais são calculados automaticamente (Quantidade x Preço Unitário).")

    wb.save(xlsx_path)
    return xlsx_path
