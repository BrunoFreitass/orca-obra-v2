"""
Permite ao usuario baixar um modelo de planilha com os precos atuais
usados pelo motor de calculo, editar os que quiser, e subir de volta --
sem precisar mexer em codigo (nem depender de mim) pra manter a tabela
de precos atualizada quando o mercado ou o SINAPI mudar.

Fora do escopo aqui: o fator regional (core/coeficientes.py) -- e um
ajuste fixo pra Roraima, unico estado atendido pelo OrçaObra hoje, nao
um preco de material/mao de obra em si.
"""
import json
import os
from datetime import UTC, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from core import coeficientes as coef
from core import paths

CAMINHO_OVERRIDES = paths.OVERRIDES_PATH


def _itens_editaveis():
    """Lista (chave, categoria, rotulo, Preco) de todo item editavel
    pela planilha. A "chave" e o elo estavel entre uma linha da
    planilha e o coeficiente certo dentro de core/coeficientes.py --
    nao depende do rotulo em portugues, que pode mudar sem quebrar o
    vinculo."""
    itens = []

    def add(chave, categoria, rotulo, preco):
        itens.append((chave, categoria, rotulo, preco))

    add("bloco_ceramico", "Material", "Bloco Cerâmico 14x19x29", coef.PRECO_BLOCO_CERAMICO)
    add("argamassa", "Material", "Argamassa AC-II (kg)", coef.PRECO_ARGAMASSA_KG)
    add("cimento", "Material", "Cimento (saco 50kg)", coef.PRECO_CIMENTO_SACO)
    add("areia", "Material", "Areia (m³)", coef.PRECO_AREIA_M3)
    add("brita", "Material", "Brita nº 1 (m³)", coef.PRECO_BRITA_M3)
    add("aco", "Material", "Aço CA-50 (kg)", coef.PRECO_ACO_RR)
    add("reboco", "Material", "Reboco (chapisco + emboço) (m²)", coef.PRECO_REBOCO_M2)
    add("impermeabilizacao", "Material", "Impermeabilização - Área Molhada (m²)", coef.PRECO_IMPERMEABILIZACAO_M2)
    add("forro_gesso", "Material", "Forro de Gesso (m²)", coef.PRECO_FORRO_GESSO_M2)
    add("rejunte", "Material", "Rejunte (kg)", coef.PRECO_REJUNTE_KG)

    grupos_por_padrao = [
        ("piso_seco", "Piso Área Seca", coef.PRECOS_PISO_SECO),
        ("piso_molhado", "Piso Área Molhada", coef.PRECOS_PISO_MOLHADO),
        ("piso_externo", "Piso Área Externa", coef.PRECOS_PISO_EXTERNO),
        ("porta_interna", "Porta Interna", coef.PRECOS_PORTA_INTERNA),
        ("porta_externa", "Porta Externa", coef.PRECOS_PORTA_EXTERNA),
        ("janela", "Janela", coef.PRECOS_JANELA),
        ("pintura", "Pintura", coef.PRECOS_PINTURA),
        ("ponto_eletrico_infra", "Ponto Elétrico - Infraestrutura", coef.PRECOS_PONTO_ELETRICO_INFRA),
        ("ponto_eletrico_acabamento", "Ponto Elétrico - Acabamento", coef.PRECOS_PONTO_ELETRICO_ACABAMENTO),
        ("ponto_hidraulico_infra", "Ponto Hidráulico - Infraestrutura", coef.PRECOS_PONTO_HIDRAULICO_INFRA),
        ("ponto_hidraulico_acabamento", "Ponto Hidráulico - Acabamento", coef.PRECOS_PONTO_HIDRAULICO_ACABAMENTO),
    ]
    for prefixo, rotulo_base, tabela in grupos_por_padrao:
        for padrao, preco in tabela.items():
            add(f"{prefixo}__{padrao}", "Material", f"{rotulo_base} ({padrao})", preco)

    for tipo_cobertura, tabela in coef.PRECOS_COBERTURA.items():
        for padrao, preco in tabela.items():
            add(f"cobertura_{tipo_cobertura}__{padrao}", "Material",
                f"Cobertura {tipo_cobertura} ({padrao})", preco)

    for servico, info in coef.MAO_DE_OBRA_POR_SERVICO.items():
        add(f"mao_de_obra__{servico}", "Mão de Obra", servico, info["preco"])

    return itens


def gerar_modelo_excel(output_path):
    """Gera um .xlsx com todos os itens editaveis e o preco vigente
    (incluindo overrides ja aplicados, se houver). O usuario baixa,
    edita a coluna 'Preço (R$)' dos itens que quiser atualizar, e sobe
    de volta em importar_tabela_excel()."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tabela de Preços"

    cabecalho = ["Chave (não editar)", "Categoria", "Item", "Preço (R$)", "Fonte atual", "Data de referência atual"]
    ws.append(cabecalho)
    for col in range(1, len(cabecalho) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    for chave, categoria, rotulo, preco_padrao in _itens_editaveis():
        preco_efetivo = obter_preco(chave, preco_padrao)
        ws.append([chave, categoria, rotulo, preco_efetivo.valor, preco_efetivo.fonte, preco_efetivo.data_ref])

    ws.column_dimensions["A"].hidden = True  # so pra religar a linha ao item certo, nao e pro usuario mexer
    for col, largura in {"B": 14, "C": 46, "D": 14, "E": 48, "F": 22}.items():
        ws.column_dimensions[col].width = largura

    wb.save(output_path)
    return output_path


def importar_tabela_excel(caminho_arquivo):
    """Le uma planilha no formato de gerar_modelo_excel() (editada ou
    nao) e retorna (precos_atualizados, avisos).

    precos_atualizados: dict {chave: novo_valor} -- so as linhas cujo
    preco realmente MUDOU em relacao ao valor vigente.
    avisos: lista de strings sobre linhas problematicas (chave
    desconhecida, preco vazio/invalido, item que sumiu da planilha).
    """
    itens_validos = {chave: preco for chave, _, _, preco in _itens_editaveis()}

    wb = load_workbook(caminho_arquivo, data_only=True)
    ws = wb.active

    precos_atualizados = {}
    avisos = []
    chaves_vistas = set()

    for linha in ws.iter_rows(min_row=2, values_only=True):
        if not linha or linha[0] is None:
            continue
        chave = str(linha[0]).strip()
        valor_bruto = linha[3] if len(linha) > 3 else None

        if chave not in itens_validos:
            avisos.append(f"Linha com chave desconhecida ignorada: '{chave}'. "
                           f"Baixe o modelo mais recente e edite a partir dele.")
            continue
        chaves_vistas.add(chave)

        if valor_bruto is None or valor_bruto == "":
            avisos.append(f"'{chave}' está com o preço em branco — mantido o valor atual.")
            continue
        try:
            novo_valor = float(valor_bruto)
        except (TypeError, ValueError):
            avisos.append(f"'{chave}' tem um valor não numérico ('{valor_bruto}') — ignorado.")
            continue
        if novo_valor <= 0:
            avisos.append(f"'{chave}' tem um preço zero ou negativo ('{novo_valor}') — ignorado.")
            continue

        preco_atual = obter_preco(chave, itens_validos[chave])
        if novo_valor != preco_atual.valor:
            precos_atualizados[chave] = novo_valor

    faltando = set(itens_validos) - chaves_vistas
    if faltando:
        exemplos = ", ".join(sorted(faltando)[:5]) + ("..." if len(faltando) > 5 else "")
        avisos.append(f"{len(faltando)} item(ns) da tabela atual não apareceram na planilha enviada "
                       f"(mantidos como estavam): {exemplos}")

    return precos_atualizados, avisos


def salvar_overrides(precos_atualizados, fonte=None, data_ref=None):
    """Persiste os precos customizados em disco (precos_customizados.json).

    fonte: string livre identificando a origem (ex: "SINAPI oficial
    (CAIXA/IBGE) - ref. 2026-07"). Se None, mantém o comportamento
    anterior (override manual via planilha).
    data_ref: "AAAA-MM" a usar como data de referência. Se None, usa
    o mês corrente (comportamento anterior, para overrides manuais
    onde não faz sentido "ref." nenhuma senão a data do upload).
    """
    data_ref = data_ref or datetime.now(tz=UTC).strftime("%Y-%m")
    dados = carregar_overrides()
    for chave, valor in precos_atualizados.items():
        entrada = {"valor": valor, "data_ref": data_ref}
        if fonte:
            entrada["fonte"] = fonte
        dados[chave] = entrada
    with open(CAMINHO_OVERRIDES, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)
    return dados


def obter_preco(chave, preco_padrao):
    """Retorna o Preco efetivo pra uma chave: o customizado (se o
    usuario tiver subido uma planilha ou importado do SINAPI alterando
    esse item), ou o padrao de core/coeficientes.py caso contrario."""
    overrides = carregar_overrides()
    if chave in overrides:
        dado = overrides[chave]
        return coef.Preco(
            valor=dado["valor"],
            fonte=dado.get("fonte", "Tabela de preços enviada pelo usuário"),
            data_ref=dado["data_ref"],
        )
    return preco_padrao


def carregar_overrides():
    """Le os precos customizados salvos em disco. Se o arquivo nao
    existir ainda, retorna vazio (usa os precos padrao de
    core/coeficientes.py em tudo)."""
    if not os.path.exists(CAMINHO_OVERRIDES):
        return {}
    with open(CAMINHO_OVERRIDES, encoding="utf-8") as f:
        return json.load(f)


def restaurar_padroes():
    """Remove todos os overrides -- volta a usar os precos padrao de
    core/coeficientes.py em tudo."""
    if os.path.exists(CAMINHO_OVERRIDES):
        os.remove(CAMINHO_OVERRIDES)


