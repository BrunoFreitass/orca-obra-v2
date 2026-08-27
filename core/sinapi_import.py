"""
Importador do SINAPI oficial (CAIXA/IBGE) para o motor de cálculo do
OrçaObra.

POR QUE O DOWNLOAD CONTINUA MANUAL: o portal da Caixa
(caixa.gov.br/poder-publico/.../sinapi) bloqueia acesso automatizado
via robots.txt. Respeitar isso significa que baixar o ZIP mensal
continua sendo um passo manual — leva ~2 minutos, uma vez por mês.
Este script cuida do resto: abrir a planilha oficial já baixada,
localizar os códigos SINAPI mapeados em core/sinapi_codigos.py,
extrair o preço vigente para Roraima e gravar como override no
OrçaObra — com a fonte e a data de referência corretas, sem digitar
nada à mão.

Esse ganho já resolve o problema real que a equipe teve na prática
(preço de cimento desatualizado no calculator.py, corrigido só depois
de perceber a olho): a partir daqui, atualizar os preços vira "baixar
o ZIP do mês e rodar um comando", em vez de lembrar de conferir cada
valor manualmente.

COMO USAR
1. Baixe e extraia o ZIP do mês em:
   https://www.caixa.gov.br/poder-publico/modernizacao-gestao/sinapi/Paginas/default.aspx
   -> Preços de Insumos e Composições -> RR -> mês mais recente ->
   versão "Não Desonerado"
2. Preencha os códigos reais em core/sinapi_codigos.py (uma vez só --
   ver instruções lá).
3. Rode, apontando para o(s) arquivo(s) .xlsx extraído(s) do ZIP:
       python -m core.sinapi_import relatorio_insumos_RR.xlsx relatorio_composicoes_RR.xlsx
4. O script imprime um resumo (o que mudou, o que não foi encontrado,
   o que ficou sem código mapeado) e pede confirmação antes de gravar.
   Use --sim para gravar sem perguntar (útil em automação futura).

O QUE ELE NÃO FAZ
Não decide se o preço extraído "faz sentido" -- isso é responsabilidade
de core/validacao.py, que já roda em cima de qualquer preço vigente
(veio do SINAPI, de override manual, ou do padrão de
core/coeficientes.py) no momento do cálculo do orçamento.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl

from core import tabela_precos as tp
from core.coeficientes import FATOR_REGIONAL_RR
from core.sinapi_codigos import MAPEAMENTO_SINAPI, UF_COLUNA_ALVOS

# core/calculator.py multiplica TODO preço por FATOR_REGIONAL_RR na hora
# de calcular o orçamento -- correto para os valores padrão de
# coeficientes.py (médias nacionais que precisam do ajuste pra virar
# preço de RR), mas os valores que vêm daqui já são a coluna RR da
# planilha oficial, ou seja, já são o preço de Roraima. Sem compensar
# isso na gravação, o cálculo aplicaria o fator regional 2x (preço 7%
# maior que o real). "aco" é a única chave que foge da regra: seu
# consumo em calculator.py NUNCA aplica o fator regional (PRECO_ACO_RR,
# o padrão dela, já era um valor de RR antes deste importador existir),
# então gravar sem compensar é o comportamento certo só pra ela.
_CHAVES_SEM_FATOR_REGIONAL = {"aco"}


def _normalizar(texto) -> str:
    return str(texto).strip().lower() if texto is not None else ""


def _achar_cabecalho(ws):
    """Procura, nas primeiras 20 linhas da planilha, a linha que
    contém uma coluna de código. Retorna (indice_linha, valores_normalizados)
    ou (None, None) se não achar -- planilhas do SINAPI têm um bloco de
    título/legenda antes do cabeçalho real, então não dá pra assumir
    linha 1."""
    for row in ws.iter_rows(min_row=1, max_row=20):
        valores = [_normalizar(c.value) for c in row]
        if any("código" in v or "codigo" in v for v in valores):
            return row[0].row, valores
    return None, None


def _achar_colunas(cabecalho_valores):
    """Localiza os índices (0-based) das colunas de código, descrição,
    unidade e preço (para a UF configurada) dentro da linha de
    cabeçalho já normalizada."""
    idx = {}
    for i, v in enumerate(cabecalho_valores):
        if ("código" in v or "codigo" in v) and "codigo" not in idx:
            idx["codigo"] = i
        elif "descri" in v and "descricao" not in idx:
            idx["descricao"] = i
        elif v.strip() in ("unid", "unidade", "und") and "unidade" not in idx:
            idx["unidade"] = i
        elif v.strip() in UF_COLUNA_ALVOS and "preco" not in idx:
            idx["preco"] = i
    return idx


def _achar_coluna_preco_multilinha(ws, linha_cab: int) -> int | None:
    """Em relatórios de composição (CSD/CCD/CSE), o cabeçalho é de duas
    linhas: o rótulo da UF (ex: 'RR') fica numa linha acima da linha
    com 'Código da Composição'/'Custo (R$)', mas alinhado na mesma
    coluna (célula âncora da mesclagem). _achar_colunas() sozinho não
    enxerga isso porque só olha a linha do 'código'. Usado como
    alternativa quando essa linha não tem a UF diretamente."""
    for linha in range(max(linha_cab - 3, 1), linha_cab):
        for cell in next(ws.iter_rows(min_row=linha, max_row=linha)):
            if cell.value is not None and _normalizar(cell.value) in UF_COLUNA_ALVOS:
                return cell.column - 1
    return None


def _extrair_mes_referencia(caminho: Path) -> str | None:
    """Tenta adivinhar o mês de referência (AAAA-MM) a partir do nome
    do arquivo baixado da Caixa. Se não conseguir, o chamador deve
    pedir --mes explicitamente."""
    m = re.search(r"(20\d{2})[-_]?(\d{2})", caminho.stem)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return None


def _sheet_e_sem_desoneracao(ws) -> bool:
    """O pacote nacional consolidado do SINAPI ('SINAPI_Referência')
    traz várias abas com o mesmo formato (ISD/ICD/ISE para insumos,
    CSD/CCD/CSE para composições) -- uma para cada regime de encargos
    (SEM/COM desoneração, SEM encargos). Sem filtrar por isso, ler
    todas as abas colidiria códigos repetidos com preços de regimes
    diferentes, e o último a ser lido venceria silenciosamente."""
    for row in ws.iter_rows(min_row=1, max_row=3, max_col=6):
        for cell in row:
            if cell.value and "sem desoneração" in _normalizar(cell.value):
                return True
    return False


def _mapa_codigos_por_descricao(wb) -> dict[tuple[str, str], str]:
    """Resolve código real de composição via descrição+unidade, usando
    a aba 'Analítico' (Relatório Analítico de Composições) quando ela
    existir no workbook.

    Necessário porque o pacote nacional consolidado do SINAPI zera a
    coluna 'Código da Composição' nas abas de custo (CSD/CCD/CSE) --
    o código de verdade só aparece na aba Analítico, na linha-resumo de
    cada composição (identificada por não ter 'Tipo Item' preenchido)."""
    for nome in wb.sheetnames:
        if "anal" not in _normalizar(nome):
            continue
        ws = wb[nome]
        linha_cab, valores_cab = _achar_cabecalho(ws)
        if linha_cab is None:
            continue
        colunas = _achar_colunas(valores_cab)
        if "codigo" not in colunas or "descricao" not in colunas or "unidade" not in colunas:
            continue
        idx_tipo = next((i for i, v in enumerate(valores_cab) if "tipo" in v), None)

        mapa = {}
        for row in ws.iter_rows(min_row=linha_cab + 1):
            tipo_cel = row[idx_tipo].value if idx_tipo is not None and idx_tipo < len(row) else None
            if tipo_cel not in (None, ""):
                continue  # linha de sub-item (insumo/mão de obra da composição), não a linha-resumo
            codigo_cel = row[colunas["codigo"]].value if colunas["codigo"] < len(row) else None
            if codigo_cel is None:
                continue
            codigo = str(codigo_cel).strip()
            if not codigo or not codigo.replace(".", "", 1).isdigit() or codigo == "0":
                continue
            descricao = row[colunas["descricao"]].value if colunas["descricao"] < len(row) else None
            unidade = row[colunas["unidade"]].value if colunas["unidade"] < len(row) else None
            mapa[(_normalizar(descricao), _normalizar(unidade))] = codigo
        return mapa
    return {}


def ler_planilha_sinapi(caminho: Path) -> dict[str, dict]:
    """Lê um arquivo .xlsx oficial do SINAPI e retorna
    {codigo_sinapi: {"preco": float, "descricao": str, "unidade": str}}
    para todos os códigos encontrados (não só os mapeados -- filtragem
    pelo mapeamento acontece depois, em importar())."""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    encontrados = {}

    sheets_sem_desoneracao = [ws for ws in wb.worksheets if _sheet_e_sem_desoneracao(ws)]
    sheets_para_ler = sheets_sem_desoneracao or wb.worksheets
    mapa_codigos = _mapa_codigos_por_descricao(wb)

    for ws in sheets_para_ler:
        linha_cab, valores_cab = _achar_cabecalho(ws)
        if linha_cab is None:
            continue
        colunas = _achar_colunas(valores_cab)
        if "preco" not in colunas:
            col_preco = _achar_coluna_preco_multilinha(ws, linha_cab)
            if col_preco is not None:
                colunas["preco"] = col_preco
        if "codigo" not in colunas or "preco" not in colunas:
            continue

        for row in ws.iter_rows(min_row=linha_cab + 1):
            codigo_cel = row[colunas["codigo"]].value if colunas["codigo"] < len(row) else None
            if codigo_cel is None:
                continue
            codigo = str(codigo_cel).strip()
            if not codigo or not codigo.replace(".", "", 1).isdigit():
                continue

            preco_cel = row[colunas["preco"]].value if colunas["preco"] < len(row) else None
            if preco_cel is None:
                continue
            try:
                preco = float(preco_cel)
            except (TypeError, ValueError):
                continue
            if preco <= 0:
                continue  # "0"/hífen = sem coleta de preço pra este estado, não um preço real

            descricao = row[colunas["descricao"]].value if "descricao" in colunas else ""
            unidade = row[colunas["unidade"]].value if "unidade" in colunas else ""

            if codigo == "0":
                codigo_real = mapa_codigos.get((_normalizar(descricao), _normalizar(unidade)))
                if codigo_real is None:
                    continue
                codigo = codigo_real

            encontrados[codigo] = {
                "preco": preco,
                "descricao": str(descricao or "").strip(),
                "unidade": str(unidade or "").strip(),
            }

    return encontrados


def importar(arquivos: list[Path], mes_referencia: str | None = None) -> tuple[dict, list[str]]:
    """Lê um ou mais arquivos oficiais do SINAPI, cruza com
    core/sinapi_codigos.py e retorna (precos_para_gravar, avisos).

    precos_para_gravar: {chave_interna: {"valor": float, "descricao": str}}
    avisos: mensagens sobre chaves sem código, códigos não encontrados
    nos arquivos, ou mudanças de unidade suspeitas.
    """
    todos_codigos: dict[str, dict] = {}
    mes_detectado = mes_referencia
    for caminho in arquivos:
        todos_codigos.update(ler_planilha_sinapi(caminho))
        if mes_detectado is None:
            mes_detectado = _extrair_mes_referencia(caminho)

    precos_para_gravar = {}
    avisos = []

    for chave, mapeamento in MAPEAMENTO_SINAPI.items():
        if mapeamento.codigo is None:
            avisos.append(f"'{chave}' ainda sem código mapeado em sinapi_codigos.py -- pulado.")
            continue

        dado = todos_codigos.get(str(mapeamento.codigo))
        if dado is None:
            avisos.append(f"'{chave}' (código {mapeamento.codigo}) não foi encontrado nos "
                           f"arquivos informados -- confira se baixou o relatório certo "
                           f"({mapeamento.tipo}).")
            continue

        unidades_incompativeis = (
            mapeamento.unidade_esperada.lower() not in dado["unidade"].lower()
            and dado["unidade"].lower() not in mapeamento.unidade_esperada.lower()
        )
        if mapeamento.unidade_esperada and dado["unidade"] and unidades_incompativeis:
            avisos.append(f"'{chave}' (código {mapeamento.codigo}): unidade no SINAPI é "
                           f"'{dado['unidade']}', esperava algo como "
                           f"'{mapeamento.unidade_esperada}' -- confira antes de confiar no valor.")

        valor_convertido = dado["preco"] * mapeamento.fator_conversao
        if mapeamento.fator_conversao != 1.0:
            avisos.append(f"'{chave}' (código {mapeamento.codigo}): preço bruto do SINAPI "
                           f"R$ {dado['preco']:.2f} convertido para R$ {valor_convertido:.2f} "
                           f"(fator ×{mapeamento.fator_conversao}) -- confira se a conversão faz sentido.")

        chave_base = chave.split("__", 1)[0]
        valor_final = round(
            valor_convertido if chave_base in _CHAVES_SEM_FATOR_REGIONAL
            else valor_convertido / FATOR_REGIONAL_RR.valor,
            2,
        )

        precos_para_gravar[chave] = {
            "valor": valor_final,
            "descricao": dado["descricao"],
        }

    if mes_detectado is None:
        avisos.append("Não consegui identificar o mês de referência pelo nome do arquivo -- "
                       "use --mes AAAA-MM para informar manualmente. Os preços não serão "
                       "gravados sem essa informação.")

    return precos_para_gravar, avisos, mes_detectado


def _resumo(precos: dict, avisos: list[str], mes_ref: str | None):
    print(f"\n{'='*70}\nResumo da importação SINAPI\n{'='*70}")
    if mes_ref:
        print(f"Mês de referência detectado: {mes_ref}\n")

    if precos:
        print(f"{len(precos)} item(ns) prontos para atualizar:\n")
        for chave, dado in precos.items():
            atual = tp.obter_preco(chave, tp.coef.Preco(0, "", "")).valor
            marcador = " (sem mudança)" if abs(atual - dado["valor"]) < 0.005 else ""
            print(f"  - {chave}: R$ {atual:.2f} -> R$ {dado['valor']:.2f}{marcador}"
                  f"  [{dado['descricao'][:60]}]")
    else:
        print("Nenhum item pronto para atualizar.")

    if avisos:
        print(f"\n{len(avisos)} aviso(s):")
        for a in avisos:
            print(f"  ! {a}")
    print()


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("arquivos", nargs="+", type=Path, help="Arquivo(s) .xlsx oficiais do SINAPI (Insumos e/ou Composições)")
    parser.add_argument("--mes", dest="mes", default=None, help="Mês de referência AAAA-MM, se não puder ser detectado pelo nome do arquivo")
    parser.add_argument("--sim", action="store_true", help="Grava sem pedir confirmação (para uso em automação)")
    args = parser.parse_args()

    for arq in args.arquivos:
        if not arq.exists():
            print(f"Arquivo não encontrado: {arq}", file=sys.stderr)
            sys.exit(1)

    precos, avisos, mes_ref = importar(args.arquivos, mes_referencia=args.mes)
    _resumo(precos, avisos, mes_ref)

    if not precos or mes_ref is None:
        print("Nada a gravar.")
        return

    if not args.sim:
        resp = input(f"Gravar {len(precos)} preço(s) como override, com fonte 'SINAPI oficial "
                      f"({mes_ref})'? [s/N] ").strip().lower()
        if resp != "s":
            print("Cancelado -- nada foi gravado.")
            return

    valores = {chave: dado["valor"] for chave, dado in precos.items()}
    tp.salvar_overrides(valores, fonte=f"SINAPI oficial (CAIXA/IBGE) - ref. {mes_ref}", data_ref=mes_ref)
    print(f"{len(valores)} preço(s) gravado(s) em {tp.CAMINHO_OVERRIDES}.")


if __name__ == "__main__":
    main()
